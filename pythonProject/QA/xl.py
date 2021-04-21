from openpyxl import *
from random import *
workbook = Workbook()
sheet = workbook.active
for i in range(1, 10):
    for x in range(1, 10):
        sheet.cell(row=i, column=x).value = randint(0, 6454)
sheet.title = "Sheet1"
print(sheet["A1"].value)
print(sheet.cell(row=2, column=2).value)
print(sheet["A:B"])
print(sheet[5])


workbook.save(filename="hello_world.xlsx")
